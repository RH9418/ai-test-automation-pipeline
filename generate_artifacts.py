import os
import sys
import yaml
import argparse
import re
import json
from datetime import datetime
import time

# --- 1. Fix the SSL Telemetry Error ---
os.environ['CREWAI_DISABLE_TELEMETRY'] = 'True'
os.environ['OTEL_SDK_DISABLED'] = 'true'

from dotenv import load_dotenv
from crewai import Agent, Task, Crew, Process, LLM
from langchain_openai import AzureChatOpenAI
from langchain_core.messages import HumanMessage, SystemMessage

# Load environment variables
load_dotenv()

API_DELAY = int(os.environ.get("RATE_LIMIT_DELAY", 15))


# --- Configuration Loader ---
def load_yaml(file_path: str) -> dict:
    with open(file_path, 'r', encoding='utf-8') as f:
        return yaml.safe_load(f)

# --- LLM Configurations ---
# LangChain wrapper (for the JSON Planner)
planner_llm = AzureChatOpenAI(
    api_key=os.environ.get("AZURE_API_KEY"),
    azure_endpoint=os.environ.get("AZURE_ENDPOINT"),
    api_version=os.environ.get("AZURE_API_VERSION"),
    azure_deployment=os.environ.get("AZURE_DEPLOYMENT_NAME"),
    temperature=0.0,
).bind(response_format={"type": "json_object"})

# CrewAI native LLM (for the Workers)
azure_llm = LLM(
    model=f"azure/{os.environ.get('AZURE_DEPLOYMENT_NAME')}", 
    api_key=os.environ.get("AZURE_API_KEY"),
    base_url=os.environ.get("AZURE_ENDPOINT"),
    api_version=os.environ.get("AZURE_API_VERSION"),
    temperature=0.0, 
    max_tokens=4000 
)

# ====================================================================
# GLOBAL HELPERS
# ====================================================================
def count_expected_actions(raw_code_content: str) -> int:
    """Accurately counts every actionable Playwright line in a snippet."""
    action_regex = re.compile(r"\.(click|fill|check|uncheck|press|hover|goto|dblclick)\(")
    expected_action_count = 0
    for line in raw_code_content.split('\n'):
        if not line.strip().startswith("#") and action_regex.search(line):
            if "close()" not in line and "chromium.launch" not in line and "new_context" not in line and "new_page" not in line:
                expected_action_count += 1
    return expected_action_count

def plan_semantic_sections(annotated_code: str) -> list:
    print("🧠 Planning semantic chunks for Artifact Generation...")
    numbered_code = "\n".join([f"{i+1}: {line}" for i, line in enumerate(annotated_code.split('\n'))])
    
    system_prompt = (
        "You are an Expert QA Architect. Divide the numbered Playwright script into cohesive, logical sections.\n"
        "Every single line MUST belong to exactly one section. No gaps.\n"
        "Return a JSON object."
    )
    message_content = (
        f"Numbered Code:\n{numbered_code}\n\n"
        "EXPECTED JSON FORMAT:\n"
        '{"sections": [{"name": "Initial Setup & Navigation", "start_line": 1, "end_line": 12}]}'
    )
    
    response = planner_llm.invoke([SystemMessage(content=system_prompt), HumanMessage(content=message_content)])
    try:
        sections = json.loads(response.content).get("sections", [])
        print(f"✅ Divided code into {len(sections)} semantic chunks.")
        return sections
    except Exception as e:
        print(f"❌ Failed to parse Planner JSON: {e}")
        return []

def extract_lines(code: str, start: int, end: int) -> str:
    lines = code.split('\n')
    return "\n".join(lines[max(0, start-1):end])

# ====================================================================
# USER'S FALLBACK BLOCK
# ====================================================================
FALLBACK_INJECTION_BLOCK = r"""
import os
import sys
import time
import re
from datetime import datetime
import pytest
from playwright.sync_api import sync_playwright, Page, expect, TimeoutError

SCREENSHOT_DIR = "Test_Screenshots"
if not os.path.exists(SCREENSHOT_DIR): os.makedirs(SCREENSHOT_DIR)
screenshot_counter = 0

def capture_annotated_screenshot(page: Page, locator, full_action_description: str):
    '''Highlights element with a thick red box and glow, then captures screenshot.'''
    global screenshot_counter
    screenshot_counter += 1
    try:
        locator.scroll_into_view_if_needed()
        locator.wait_for(state='visible', timeout=5000)
        box = locator.bounding_box()
        if not box: return
            
        js_description = full_action_description.replace("'", "\\'")
        page.evaluate(f'''(params) => {{
            const {{ box, desc }} = params;
            const div = document.createElement('div');
            div.id = 'ge-spotlight-box'; div.style.position = 'absolute';
            div.style.left = `${{box.x}}px`; div.style.top = `${{box.y}}px`;
            div.style.width = `${{box.width}}px`; div.style.height = `${{box.height}}px`;
            div.style.border = '5px solid #FF0000'; div.style.boxShadow = '0 0 15px 5px rgba(255, 0, 0, 0.7)';
            div.style.boxSizing = 'border-box'; div.style.zIndex = '2147483647'; div.style.pointerEvents = 'none';
            
            const label = document.createElement('div');
            label.id = 'ge-spotlight-label'; label.textContent = desc; label.style.position = 'absolute';
            label.style.left = `${{box.x}}px`; label.style.top = `${{box.y - 40 > 0 ? box.y - 40 : box.y + box.height + 10}}px`;
            label.style.backgroundColor = '#FF0000'; label.style.color = '#FFFFFF'; label.style.padding = '8px 12px';
            label.style.fontSize = '16px'; label.style.fontWeight = 'bold'; label.style.borderRadius = '4px';
            label.style.zIndex = '2147483647';
            document.body.appendChild(div); document.body.appendChild(label);
        }}''', {'box': box, 'desc': js_description})
        
        time.sleep(0.2)
        timestamp = datetime.now().strftime("%H-%M-%S")
        safe_filename = re.sub(r'[^a-z0-9]', '_', full_action_description.lower())[:40]
        filename = f"{timestamp}_{screenshot_counter:02d}_{safe_filename}.png"
        path = os.path.join(SCREENSHOT_DIR, filename)
        
        page.screenshot(path=path, full_page=False)
        print(f"   └── 📸 Screenshot saved: {path}")
        page.evaluate("() => { document.getElementById('ge-spotlight-box')?.remove(); document.getElementById('ge-spotlight-label')?.remove(); }")
    except Exception as e: print(f"   └── ⚠️ Screenshot Error: {e}")

def safe_action(page: Page, locator, action_name: str, description: str, *action_args):
    '''Performs action with spotlight screenshots and manual fallbacks.'''
    full_desc = f"{action_name.capitalize()}: {description}"
    if action_name == 'fill': full_desc += f" with '{action_args[0] if action_args else ''}'"
        
    try:
        if action_name == 'fill':
            print(f"\n⏸️ PAUSING FOR INPUT: About to fill '{description}'.")
            input("   └── Perform input manually in browser, then PRESS [ENTER] to continue...")
            page.wait_for_load_state('networkidle', timeout=5000)
            return

        if action_name in ['click', 'dblclick', 'check', 'uncheck', 'hover']:
            try: locator.hover(timeout=2000); time.sleep(0.3)
            except: pass
            
        if action_name == 'close':
            try: locator.close()
            except: pass
            print(f"✅ SUCCESS: {description} (Teardown handled by Pytest)")
            return

        if locator != page:
            capture_annotated_screenshot(page, locator, full_desc)
            action_func = getattr(locator, action_name)
            action_func(*action_args)
        else:
            if action_name == 'goto': page.goto(*action_args)
                
        print(f"✅ SUCCESS: {description}")
    except Exception as e:
        print(f"❌ ERROR: Failed {action_name} on '{description}'.")
        while True:
            print("\n" + "="*80 + "\n ACTION REQUIRED: Script Error\n" + "="*80)
            print(f" Failed: {full_desc}")
            choice = input(" Did you perform this manually? (y/n): ").lower().strip()
            if choice == 'y': break
            elif choice == 'n': sys.exit(0)
    finally:
        try: page.wait_for_load_state('networkidle', timeout=3000)
        except: time.sleep(1)
"""

# ====================================================================
# PHASE 1: Generate Feature Documentation (CHUNKED)
# ====================================================================
def process_annotated_script_to_docs(annotated_code: str, feature_name: str, tab_name: str, sections: list) -> str:
    print(f"\n--- [PHASE 1] Generating Markdown Documentation (Chunk-by-Chunk) ---")
    agents_config = load_yaml('configs/agents.yaml')['agents']
    
    # 🔴 FIX: Upgraded to strict 6-column format matching the Excel sheet
    master_md = (
        f"# Feature: {feature_name}\n"
        f"**Tab Location:** {tab_name}\n\n"
        f"## Detailed User Interaction Flows\n\n"
        f"| Use Case ID | Test Case Description | Exact Locator | Pass/Fail | Notes | Issues |\n"
        f"|:---|:---|:---|:---|:---|:---|\n"
    )
    
    prev_context = "None. This is the first section."
    
    for i, sec in enumerate(sections):
        chunk_code = extract_lines(annotated_code, sec['start_line'], sec['end_line'])
        expected_count = count_expected_actions(chunk_code)
        
        if expected_count == 0:
            continue 
            
        print(f"   └── 📝 UAT Section {i+1}/{len(sections)}: {sec['name']} (Expected Actions: {expected_count})")
        section_prefix = chr(65 + i) 
        
        for attempt in range(3):
            expander_agent = Agent(
                role=agents_config['detail_expander_agent']['role'], 
                goal=agents_config['detail_expander_agent']['goal'], 
                backstory=agents_config['detail_expander_agent']['backstory'], 
                allow_delegation=False, 
                llm=azure_llm
            )
            
            task = Task(
                description=(
                    "You are a Strict Markdown Table Formatter and an Expert Technical Writer. Convert the provided Code Snippet into Markdown table rows.\n\n"
                    f"CRITICAL RULES FOR NO DATA LOSS:\n"
                    f"1. You MUST generate EXACTLY {expected_count} rows containing locators. Every single Playwright action MUST get its own row.\n"
                    "2. DO NOT include the markdown table header. Output ONLY the table body rows.\n"
                    f"3. Start the section with a header row exactly like this: `| | **--- {sec['name'].upper()} ---** | | | | |`\n"
                    f"4. Use the prefix '{section_prefix}' for the Use Case IDs (e.g., {section_prefix}.1, {section_prefix}.2).\n\n"
                    "GUIDEBOOK FORMATTING RULES:\n"
                    "1. The 'Test Case Description' column MUST be highly readable for manual testers. Write it like an instruction manual or guidebook.\n"
                    "2. Instead of robotic fragments (e.g., 'Click button'), write clear, descriptive steps explaining exactly what to do and where to look (e.g., 'From the filter panel on the left, click on the \"Apply Filters\" button to execute the search.').\n"
                    "3. You MUST use exactly 6 columns: `| Use Case ID | Test Case Description | Exact Locator | Pass/Fail | Notes | Issues |`\n"
                    "4. Column 3 MUST contain the EXACT raw Playwright code line wrapped in backticks (e.g. `page.locator(...).click()`).\n"
                    "5. Columns 4, 5, and 6 must be empty cells (just a space between pipes: `| | | |`).\n\n"
                    f"CONTEXT FROM PREVIOUS SECTION:\n{prev_context}\n\n"
                    f"CODE SNIPPET TO CONVERT:\n{chunk_code}"
                ),
                expected_output="Raw Markdown table rows only, strictly adhering to the 6-column format.",
                agent=expander_agent
            )
            crew = Crew(agents=[expander_agent], tasks=[task], process=Process.sequential)
            
            try:
                result = str(crew.kickoff()).strip()
                md_action_count = len(re.findall(r"\.(click|fill|check|uncheck|press|hover|goto|dblclick)\(", result))
                
                if md_action_count >= expected_count:
                    master_md += result + "\n"
                    prev_context = result
                    print(f"       ✅ Passed. Waiting {API_DELAY}s for API Rate Limit...")
                    time.sleep(API_DELAY)
                    break
                else:
                    print(f"       ❌ Dropped lines ({md_action_count}/{expected_count}). Retrying after {API_DELAY}s...")
                    time.sleep(API_DELAY)
            except Exception as e:
                print(f"       ⚠️ API Error: {e}. Retrying after {API_DELAY}s...")
                time.sleep(API_DELAY)
                
    return master_md



# ====================================================================
# PHASE 2: Generate Pytest Scripts (CHUNKED)
# ====================================================================
def process_docs_to_pytest(uat_content: str, annotated_code: str, output_pytest_path: str, sections: list):
    print(f"\n--- [PHASE 2] Generating Smart Pytest Scripts (Chunk-by-Chunk) ---")
    agents_config = load_yaml('configs/agents.yaml')['agents']

    master_pytest = FALLBACK_INJECTION_BLOCK + "\n\n"
    prev_context = "None. This is the first test function."

    for i, sec in enumerate(sections):
        chunk_code = extract_lines(annotated_code, sec['start_line'], sec['end_line'])
        expected_count = count_expected_actions(chunk_code)
        
        if expected_count == 0:
            continue
            
        print(f"   └── 📝 Pytest Function {i+1}/{len(sections)}: {sec['name']} (Expected Actions: {expected_count})")
        func_name = re.sub(r'[^a-zA-Z0-9_]', '_', sec['name'].lower())
        
        for attempt in range(3):
            # 🔴 FIX: Instantiate fresh agent INSIDE the loop to wipe memory
            automation_agent = Agent(role=agents_config['automation_agent']['role'], goal=agents_config['automation_agent']['goal'], backstory=agents_config['automation_agent']['backstory'], allow_delegation=False, llm=azure_llm)

            task = Task(
                description=(
                    "You are an Expert SDET. Convert the provided Code Snippet into a SINGLE Pytest function.\n"
                    f"1. Name the function exactly: `test_{func_name}(shared_page: Page):`\n"
                    f"2. Add the Pytest ordering decorator directly above the function: `@pytest.mark.order({i+1})`\n"
                    f"3. You MUST wrap every interaction in `safe_action()`. Generate EXACTLY {expected_count} `safe_action` calls.\n"
                    "4. Do NOT output boilerplate or imports. Output ONLY the decorator and the single function.\n"
                    "5. Reference the provided UAT Document to generate the descriptive string for safe_action.\n\n"
                    f"CONTEXT FROM PREVIOUS FUNCTION:\n{prev_context}\n\n"
                    f"MASTER UAT DOCUMENT (For Reference):\n{uat_content}\n\n"
                    f"CODE SNIPPET TO CONVERT:\n{chunk_code}"
                ),
                expected_output="A single python function block starting with `@pytest.mark.order`.",
                agent=automation_agent
            )
            crew = Crew(agents=[automation_agent], tasks=[task], process=Process.sequential)
            
            try:
                result = str(crew.kickoff()).strip()
                cleaned_code = result
                if cleaned_code.startswith("```python"): cleaned_code = cleaned_code[9:]
                elif cleaned_code.startswith("```"): cleaned_code = cleaned_code[3:]
                if cleaned_code.endswith("```"): cleaned_code = cleaned_code[:-3]
                cleaned_code = cleaned_code.strip()

                # 🔴 FIX: Relaxed count criteria slightly in case LLM names the page argument differently
                actual_action_count = cleaned_code.count("safe_action(")
                
                if actual_action_count >= expected_count:
                    master_pytest += cleaned_code + "\n\n"
                    prev_context = cleaned_code
                    print(f"       ✅ Passed. Waiting {API_DELAY}s for API Rate Limit...")
                    time.sleep(API_DELAY)
                    break
                else:
                    print(f"       ❌ Dropped lines ({actual_action_count}/{expected_count}). Retrying after {API_DELAY}s...")
                    time.sleep(API_DELAY)
            except Exception as e:
                print(f"       ⚠️ API Error: {e}. Retrying after {API_DELAY}s...")
                time.sleep(API_DELAY)

    with open(output_pytest_path, 'w', encoding='utf-8') as f:
        f.write(master_pytest)
    print(f"\n✅ Successfully generated final Pytest script: {output_pytest_path}")


import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill

# ====================================================================
# PHASE 3: Generate UAT Excel Sheets (WITH 0 DATA LOSS CHECK)
# ====================================================================
def run_step_3_generate_excel_uat(docs_dir: str, excel_dir: str):
    print("\n--- [PHASE 3] Converting Markdown to Advanced UAT Excel ---")
    if not os.path.exists(docs_dir):
        print(f"❌ Error: Directory '{docs_dir}' not found.")
        return
        
    processed_count = 0
    
    for md_file in os.listdir(docs_dir):
        if not md_file.endswith(".md"): continue
        
        md_file_path = os.path.join(docs_dir, md_file)
        print(f"\n   └── 📄 Processing: {md_file}")
        
        with open(md_file_path, "r", encoding="utf-8") as f:
            lines = f.readlines()
        
        rows = []
        current_feature = ""
        raw_table_row_count = 0 # Used for the 0 Data Loss Check
        
        for line in lines:
            line = line.strip()
            
            # Ignore markdown code block wrappers
            if line.startswith("```"):
                continue
                
            # Capture the Main Feature Title
            if line.startswith("# Feature:"):
                current_feature = line.replace("# Feature:", "").strip()
                rows.append({
                    "Use Case ID": "", 
                    "Test Case Description": f"--- {current_feature.upper()} ---", 
                    "Exact Locator": "", "Pass/Fail": "", "Notes": "", "Issues": ""
                })
                continue
                
            # Skip Markdown table headers and alignment rows
            if line.startswith("|") and "Use Case ID" in line:
                continue 
            if line.startswith("|") and ":---" in line:
                continue 
                
            # Parse Data Rows
            if line.startswith("|"):
                raw_table_row_count += 1
                parts = [p.strip() for p in line.split("|")[1:-1]]
                
                if len(parts) >= 6:
                    use_case = parts[0]
                    desc = parts[1].replace("<br>", "\n").replace("<br/>", "\n")
                    
                    # Clean up bolding asterisks from section headers
                    if desc.startswith("**---") and desc.endswith("---**"):
                        desc = desc.replace("**", "")
                        
                    rows.append({
                        "Use Case ID": use_case,
                        "Test Case Description": desc,
                        "Exact Locator": parts[2],
                        "Pass/Fail": parts[3],
                        "Notes": parts[4],
                        "Issues": parts[5]
                    })

        if rows:
            excel_filename = md_file.replace(".md", ".xlsx")
            excel_path = os.path.join(excel_dir, excel_filename)
            df = pd.DataFrame(rows)
            
            # --- 0 DATA LOSS VERIFICATION ---
            # Subtract 1 for the main feature header we manually injected
            parsed_rows = len(rows) - (1 if current_feature else 0)
            if parsed_rows != raw_table_row_count:
                print(f"       ❌ WARNING: Possible data loss! Raw Markdown rows: {raw_table_row_count}, Excel Rows: {parsed_rows}")
            else:
                print(f"       ✅ Verified: 0 Data Loss. Parsed exactly {raw_table_row_count} rows.")

            # --- EXCEL STYLING ---
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='UAT Summary')
                worksheet = writer.sheets['UAT Summary']
                
                header_font = Font(bold=True, color="FFFFFF", size=11)
                header_fill = PatternFill(start_color="4F81BD", fill_type="solid")
                part_font = Font(bold=True, color="FFFFFF", size=12)
                part_fill = PatternFill(start_color="1F4E78", fill_type="solid")
                
                wrap_alignment = Alignment(wrap_text=True, vertical="top")
                center_alignment = Alignment(horizontal="center", vertical="center")

                # Style Headers
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_alignment
                
                # Adjust Column Widths
                worksheet.column_dimensions['A'].width = 15
                worksheet.column_dimensions['B'].width = 65
                worksheet.column_dimensions['C'].width = 50
                worksheet.column_dimensions['D'].width = 12
                worksheet.column_dimensions['E'].width = 25
                worksheet.column_dimensions['F'].width = 25

                # Style Rows and Merge Headers
                for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), 2):
                    use_case_cell, description_cell, locator_cell = row[0], row[1], row[2]
                    desc_val = str(description_cell.value)
                    
                    if desc_val.startswith("--- ") and desc_val.endswith(" ---"):
                        # This is a Section Header - Merge it across the table
                        worksheet.merge_cells(f'A{row_idx}:F{row_idx}')
                        merged_cell = worksheet.cell(row=row_idx, column=1)
                        merged_cell.value = desc_val
                        merged_cell.font = part_font
                        merged_cell.fill = part_fill
                        merged_cell.alignment = center_alignment
                    else:
                        # Standard Row
                        use_case_cell.alignment = center_alignment
                        description_cell.alignment = wrap_alignment
                        locator_cell.alignment = wrap_alignment
                        row[3].alignment = center_alignment
                        row[4].alignment = wrap_alignment
                        row[5].alignment = wrap_alignment

            print(f"       ✅ Generated Styled Excel: {excel_path}")
            processed_count += 1
            
    print(f"\n--- Excel Conversion Complete! Created {processed_count} files. ---")


# --- Execution Entry Point ---
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Orchestrate CrewAI generation via Semantic Chunking.")
    parser.add_argument('--step', type=int, choices=[1, 2, 3], help="Phase to run (1: Docs, 2: Pytest, 3: Excel). If omitted, runs all.")
    parser.add_argument('--files', nargs='*', help="Specific annotated script names to process.")
    args = parser.parse_args()


    annotated_logs_dir = "Annotated_Logs"
    output_docs_dir = "Generated_Documentation"
    output_pytest_dir = "Generated_Scripts"
    output_excel_dir = "UAT_Excel_Reports" 
    
    os.makedirs(output_docs_dir, exist_ok=True)
    os.makedirs(output_pytest_dir, exist_ok=True)
    os.makedirs(output_excel_dir, exist_ok=True) # Ensure it exists
    
    if args.files:
        available_files = os.listdir(annotated_logs_dir)
        scripts_to_process = [f for f in args.files if f in available_files]
    else:
        scripts_to_process = [f for f in os.listdir(annotated_logs_dir) if f.endswith(".py")]
    
    if not scripts_to_process:
        sys.exit("No annotated scripts found to process.")

    run_step_1 = args.step is None or args.step == 1
    run_step_2 = args.step is None or args.step == 2
    run_step_3 = args.step is None or args.step == 3

    for script_name in scripts_to_process:
        print(f"\n{'='*60}\n🚀 Processing: {script_name}\n{'='*60}")
        script_path = os.path.join(annotated_logs_dir, script_name)
        feature_name_derived = script_name.replace(".py", "").replace("_", " ").title()
        
        doc_output_path = os.path.join(output_docs_dir, f"{feature_name_derived}_UAT.md")
        pytest_output_path = os.path.join(output_pytest_dir, f"test_{script_name}")
        
        with open(script_path, 'r', encoding='utf-8') as f:
            annotated_code = f.read()
            
        # 1. Plan the Semantic Chunks ONCE for both phases
        sections = plan_semantic_sections(annotated_code)
        if not sections:
            print("Failed to chunk file. Skipping.")
            continue

        # --- PHASE 1 (Docs) ---
        if run_step_1:
            md_result = process_annotated_script_to_docs(annotated_code, feature_name_derived, "Main Workspace", sections)
            with open(doc_output_path, 'w', encoding='utf-8') as f:
                f.write(md_result)
            print(f"✅ Saved Verified Docs: {doc_output_path}")
        
        # --- PHASE 2 (Pytest) ---
        if run_step_2:
            if not os.path.exists(doc_output_path):
                print(f"❌ Error: Required UAT document missing.")
            else:
                with open(doc_output_path, 'r', encoding='utf-8') as f: uat_content = f.read()
                process_docs_to_pytest(uat_content, annotated_code, pytest_output_path, sections)
        if run_step_3:
            run_step_3_generate_excel_uat(output_docs_dir, output_excel_dir)
