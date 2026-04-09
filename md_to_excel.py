import os
import re
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill

def convert_md_to_excel(docs_dir: str, excel_dir: str):
    print(f"\n--- Converting Markdown to Advanced UAT Excel ---")
    
    if not os.path.exists(docs_dir):
        print(f"❌ Error: Directory '{docs_dir}' not found.")
        return
        
    os.makedirs(excel_dir, exist_ok=True)
    processed_count = 0
    
    for md_file in os.listdir(docs_dir):
        if not md_file.endswith(".md"): 
            continue
        
        md_file_path = os.path.join(docs_dir, md_file)
        print(f"\n   └── 📄 Processing: {md_file}")
        
        with open(md_file_path, "r", encoding="utf-8") as f:
            lines = f.readlines()
        
        rows = []
        current_feature = ""
        raw_table_row_count = 0 # Used for the 0 Data Loss Check
        
        for line in lines:
            line = line.strip()
            
            # Ignore markdown code block wrappers
            if line.startswith("```"):
                continue
                
            # Capture the Main Feature Title
            if line.startswith("# Feature:"):
                current_feature = line.replace("# Feature:", "").strip()
                rows.append({
                    "Use Case ID": "", 
                    "Test Case Description": f"--- {current_feature.upper()} ---", 
                    "Exact Locator": "", "Pass/Fail": "", "Notes": "", "Issues": ""
                })
                continue
                
            # Skip Markdown table headers and alignment rows
            if line.startswith("|") and "Use Case ID" in line:
                continue 
            if line.startswith("|") and ":---" in line:
                continue 
                
            # Parse Data and Section Header Rows
            if line.startswith("|"):
                raw_table_row_count += 1
                
                # Split by pipe, ignoring the first and last empty strings from the edges
                parts = [p.strip() for p in line.split("|")[1:-1]]
                
                if len(parts) >= 6:
                    use_case = parts[0]
                    desc = parts[1].replace("<br>", "\n").replace("<br/>", "\n")
                    
                    # Clean up bolding asterisks from section headers if present
                    if desc.startswith("**---") and desc.endswith("---**"):
                        desc = desc.replace("**", "")
                        
                    rows.append({
                        "Use Case ID": use_case,
                        "Test Case Description": desc,
                        "Exact Locator": parts[2],
                        "Pass/Fail": parts[3],
                        "Notes": parts[4],
                        "Issues": parts[5]
                    })

        if rows:
            excel_filename = md_file.replace(".md", ".xlsx")
            excel_path = os.path.join(excel_dir, excel_filename)
            df = pd.DataFrame(rows)
            
            # --- 0 DATA LOSS VERIFICATION ---
            # Subtract 1 for the main feature header we manually injected at the top
            parsed_rows = len(rows) - (1 if current_feature else 0)
            if parsed_rows != raw_table_row_count:
                print(f"       ❌ WARNING: Possible data loss! Raw Markdown rows: {raw_table_row_count}, Parsed Rows: {parsed_rows}")
            else:
                print(f"       ✅ Verified: 0 Data Loss. Successfully parsed exactly {raw_table_row_count} rows.")

            # --- EXCEL STYLING ---
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='UAT Summary')
                worksheet = writer.sheets['UAT Summary']
                
                header_font = Font(bold=True, color="FFFFFF", size=11)
                header_fill = PatternFill(start_color="4F81BD", fill_type="solid")
                part_font = Font(bold=True, color="FFFFFF", size=12)
                part_fill = PatternFill(start_color="1F4E78", fill_type="solid")
                
                wrap_alignment = Alignment(wrap_text=True, vertical="top")
                center_alignment = Alignment(horizontal="center", vertical="center")

                # Style Headers
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_alignment
                
                # Adjust Column Widths
                worksheet.column_dimensions['A'].width = 15
                worksheet.column_dimensions['B'].width = 65
                worksheet.column_dimensions['C'].width = 50
                worksheet.column_dimensions['D'].width = 12
                worksheet.column_dimensions['E'].width = 25
                worksheet.column_dimensions['F'].width = 25

                # Style Rows and Merge Headers
                for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), 2):
                    use_case_cell, description_cell, locator_cell = row[0], row[1], row[2]
                    desc_val = str(description_cell.value)
                    
                    if desc_val.startswith("--- ") and desc_val.endswith(" ---"):
                        # This is a Section Header - Merge it across the table
                        worksheet.merge_cells(f'A{row_idx}:F{row_idx}')
                        merged_cell = worksheet.cell(row=row_idx, column=1)
                        merged_cell.value = desc_val
                        merged_cell.font = part_font
                        merged_cell.fill = part_fill
                        merged_cell.alignment = center_alignment
                    else:
                        # Standard Data Row
                        use_case_cell.alignment = center_alignment
                        description_cell.alignment = wrap_alignment
                        locator_cell.alignment = wrap_alignment
                        row[3].alignment = center_alignment
                        row[4].alignment = wrap_alignment
                        row[5].alignment = wrap_alignment

            print(f"       ✅ Generated Styled Excel: {excel_path}")
            processed_count += 1
            
    print(f"\n--- Excel Conversion Complete! Created {processed_count} files. ---")


if __name__ == "__main__":
    docs_dir = "Generated_Documentation"
    excel_dir = "UAT_Excel_Reports"
    
    # Run the conversion
    convert_md_to_excel(docs_dir, excel_dir)
